from openpyxl import load_workbook
import datetime


class ExcelWriter(object):

    def __init__(self, transactions, workbook):
        self.path = workbook
        self.transactions = transactions
        self.workbook = load_workbook(workbook)
        self.ws = self.workbook['Data']

    def write_to_file(self):
        current_row = self._find_max_row()
        for transaction in self.transactions:
            parsed_date = datetime.datetime.strptime(transaction.transaction_date, '%b %d %Y')
            self.ws['A'+str(current_row)] = parsed_date
            self.ws['B'+str(current_row)] = transaction.description
            self.ws['C'+str(current_row)] = float(transaction.amount)
            self.ws['C'+str(current_row)].number_format = u'"$"#,##0.00'
            self.ws['D'+str(current_row)] = transaction.account
            self.ws['I' + str(current_row)] = transaction.sub_category
            self.ws['J' + str(current_row)] = transaction.main_category
            current_row = current_row + 1
        self.workbook.save(self.path)

    def _sort_by_date(self):
        pass

    def _find_max_row(self):
        for max_row, row in enumerate(self.ws, 1):
            if all(c.value is None for c in row):
                return max_row
        return max_row

